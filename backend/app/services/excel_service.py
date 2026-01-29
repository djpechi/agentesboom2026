# /backend/app/services/excel_service.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Any
from io import BytesIO


def generate_excel(
    account_name: str,
    company_website: str | None,
    stage_outputs: dict[int, dict[str, Any]]
) -> bytes:
    """
    Generate Excel report from stage outputs
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create Summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    _create_summary_sheet(summary_sheet, account_name, company_website, stage_outputs)

    # Create Stage 1 sheets if available
    if 1 in stage_outputs:
        s1_data = stage_outputs[1]
        
        # Scaling Up Table Sheet
        if 'scalingUpTable' in s1_data:
            su_sheet = wb.create_sheet("1. Scaling Up Table")
            _create_scaling_up_sheet(su_sheet, s1_data['scalingUpTable'])
            
        # Buyer Persona Sheet
        if 'buyerPersona' in s1_data:
            bp_sheet = wb.create_sheet("1. Buyer Persona")
            _create_buyer_persona_sheet(bp_sheet, s1_data['buyerPersona'])

    # Create Stage 2 sheet if available
    if 2 in stage_outputs:
        stage2_sheet = wb.create_sheet("2. Customer Journey")
        _create_stage2_sheet(stage2_sheet, stage_outputs[2])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def _create_summary_sheet(
    sheet,
    account_name: str,
    company_website: str | None,
    stage_outputs: dict[int, dict[str, Any]]
):
    """Create summary sheet with account info"""
    sheet['A1'] = 'BOOMS Platform - Marketing Onboarding Report'
    sheet['A1'].font = Font(bold=True, size=16, color="2563eb")
    sheet.merge_cells('A1:D1')

    row = 3
    sheet[f'A{row}'] = 'Client:'
    sheet[f'B{row}'] = account_name
    sheet[f'A{row}'].font = Font(bold=True)

    row += 1
    sheet[f'A{row}'] = 'Website:'
    sheet[f'B{row}'] = company_website or 'N/A'
    sheet[f'A{row}'].font = Font(bold=True)

    row += 1
    sheet[f'A{row}'] = 'Generated:'
    sheet[f'B{row}'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    sheet[f'A{row}'].font = Font(bold=True)

    row += 2
    sheet[f'A{row}'] = 'Completed Stages'
    sheet[f'A{row}'].font = Font(bold=True, size=14)

    for stage_num in sorted(stage_outputs.keys()):
        row += 1
        stage_names = {
            1: "Stage 1 - BOOMS (Buyer Persona Architect)",
            2: "Stage 2 - Journey (Customer Journey Mapping)"
        }
        sheet[f'A{row}'] = stage_names.get(stage_num, f"Stage {stage_num}")

    for col in ['A', 'B', 'C', 'D']:
        sheet.column_dimensions[col].width = 30


def _create_scaling_up_sheet(sheet, table_data: list[dict]):
    """Create Scaling Up Table sheet"""
    sheet['A1'] = 'Scaling Up Table - Buyer Persona Criteria'
    sheet['A1'].font = Font(bold=True, size=14, color="1e40af")
    
    headers = ["Criterion", "Super Green", "Green", "Yellow", "Red", "Not Eligible"]
    header_colors = ["FFFFFF", "22c55e", "86efac", "fde047", "f87171", "475569"]
    
    for i, header in enumerate(headers):
        cell = sheet.cell(row=3, column=i+1)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF" if i != 3 else "000000")
        cell.fill = PatternFill(start_color=header_colors[i], end_color=header_colors[i], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 4
    for item in table_data:
        sheet.cell(row=row, column=1, value=item.get('criterion', ''))
        sheet.cell(row=row, column=2, value=item.get('superGreen', ''))
        sheet.cell(row=row, column=3, value=item.get('green', ''))
        sheet.cell(row=row, column=4, value=item.get('yellow', ''))
        sheet.cell(row=row, column=5, value=item.get('red', ''))
        sheet.cell(row=row, column=6, value=item.get('notEligible', ''))
        
        for c in range(1, 7):
            sheet.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Adjust widths
    sheet.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        sheet.column_dimensions[col].width = 35


def _create_buyer_persona_sheet(sheet, persona_data: dict):
    """Create Buyer Persona narrative sheet"""
    sheet['A1'] = f"Buyer Persona: {persona_data.get('name', 'N/A')}"
    sheet['A1'].font = Font(bold=True, size=16, color="1e40af")
    
    row = 3
    sections = [
        ("Demographics", 'demographics'),
        ("Professional Context", 'professionalContext'),
        ("Goals", 'goals'),
        ("Challenges", 'challenges'),
        ("Behaviors", 'behaviors'),
        ("Narrative", 'narrative')
    ]
    
    for title, key in sections:
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        sheet[f'A{row}'].fill = PatternFill(start_color="eff6ff", end_color="eff6ff", fill_type="solid")
        row += 1
        
        val = persona_data.get(key, 'N/A')
        if isinstance(val, list):
            for item in val:
                sheet[f'A{row}'] = f"• {item}"
                row += 1
        else:
            sheet[f'A{row}'] = val
            sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    sheet.column_dimensions['A'].width = 100


def _create_stage2_sheet(sheet, stage_data: dict[str, Any]):
    """Create Stage 2 (Journey) sheet"""
    sheet['A1'] = 'Stage 2: Journey - Customer Journey Mapping'
    sheet['A1'].font = Font(bold=True, size=14, color="1e40af")
    
    row = 3
    journey_stages = stage_data.get('stages', [])
    for s in journey_stages:
        sheet[f'A{row}'] = s.get('name', 'N/A')
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        sheet[f'A{row}'].fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        row += 1
        
        for key, title in [('touchpoints', 'Touchpoints'), ('pain_points', 'Pain Points'), ('opportunities', 'Opportunities')]:
            sheet[f'A{row}'] = title + ":"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            for item in s.get(key, []):
                sheet[f'B{row}'] = f"• {item}"
                row += 1
        row += 1

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 75
